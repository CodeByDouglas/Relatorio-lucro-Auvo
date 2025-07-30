from app.models.user import User
from app.models.tarefas import Tarefas
from app.models.servicos import Servicos
from app.models.tipos_de_tarefas import Tipos_de_tarefas
import os
import tempfile
import openpyxl
from openpyxl.styles import Border, Side, PatternFill

def extrair_dados_servico(user_id):
    """
    Função que extrai e processa os dados para o relatório de serviços
    """
    # Buscar dados das tarefas
    tarefas_obj = Tarefas.query.filter_by(user_id=user_id).first()
    tarefas = tarefas_obj.json_lista_tarefas if tarefas_obj and tarefas_obj.json_lista_tarefas else []

    # Filtrar apenas tarefas com faturamento de serviços diferente de zero
    tarefas_servicos = []
    for tarefa in tarefas:
        faturamento_servicos = tarefa.get('faturamento-servicos', 0)
        if faturamento_servicos and faturamento_servicos != 0:
            tarefas_servicos.append(tarefa)

    # Buscar serviços cadastrados
    servicos_obj = Servicos.query.filter_by(user_id=user_id).first()
    servicos_cadastrados = servicos_obj.json_lista_servicos if servicos_obj and servicos_obj.json_lista_servicos else []

    # Montar lista de tarefas com os campos desejados (apenas serviços)
    tarefas_base = []
    for tarefa in tarefas_servicos:
        tarefas_base.append({
            'Código da Tarefa': tarefa.get('id-da-tarefa'),
            'Data': tarefa.get('data-da-tarefa', '')[:10],
            'Cliente': tarefa.get('nome-do-cliente'),
            'Colaborador': tarefa.get('id-do-colaborador'),
            'Tipo de Tarefa': tarefa.get('tipo-da-tarefa'),
            'Serviços': tarefa.get('serviços', []),
            'Faturamento': tarefa.get('faturamento-servicos', 0),
            'Lucro': tarefa.get('lucro-servicos', 0)
        })

    # Criar dicionário de id->nome para serviços cadastrados
    servicos_map = {}
    for s in servicos_cadastrados:
        if isinstance(s, dict) and 'id-servico' in s and 'nome-do-servico' in s:
            servicos_map[str(s['id-servico'])] = s['nome-do-servico']

    # Montar lista de tarefas com nomes de serviços
    tarefas_com_servicos = []
    for tarefa in tarefas_base:
        nomes_servicos = [servicos_map.get(str(sid), sid) for sid in tarefa['Serviços']]
        tarefa_nome = tarefa.copy()
        tarefa_nome['Serviços'] = nomes_servicos
        tarefas_com_servicos.append(tarefa_nome)

    # Buscar tipos de tarefa cadastrados
    tipos_obj = Tipos_de_tarefas.query.filter_by(user_id=user_id).first()
    tipos_cadastrados = tipos_obj.json_lista_tipos_de_tarefas if tipos_obj and tipos_obj.json_lista_tipos_de_tarefas else []

    # Criar dicionário de id->nome para tipos de tarefa cadastrados
    tipos_map = {}
    for t in tipos_cadastrados:
        if isinstance(t, dict) and 'id-tipo-de-tarefa' in t and 'nome-do-tipo-de-tarefa' in t:
            tipos_map[str(t['id-tipo-de-tarefa'])] = t['nome-do-tipo-de-tarefa']

    # Montar lista final de tarefas com nomes de serviços e tipos de tarefa
    tarefas_processadas = []
    for tarefa in tarefas_com_servicos:
        tipo_nome = tipos_map.get(str(tarefa['Tipo de Tarefa']), tarefa['Tipo de Tarefa'])
        tarefa_nome = tarefa.copy()
        tarefa_nome['Tipo de Tarefa'] = tipo_nome
        tarefas_processadas.append(tarefa_nome)

    return tarefas_processadas


def gerar_planilha_excel_servico(user_id):
    """
    Função que gera a planilha Excel para o relatório de serviços
    """
    # Extrair e processar os dados
    tarefas_processadas = extrair_dados_servico(user_id)
    # Gerar planilha Excel usando o modelo de serviços
    modelo_path = os.path.join(os.path.dirname(__file__), 'modelos', 'Relatorio_de_Lucro_servico.xlsx')
    wb = openpyxl.load_workbook(modelo_path)
    ws = wb.active

    # Limpar linhas antigas de dados (exceto cabeçalho)
    max_data_row = ws.max_row
    if max_data_row > 2:
        ws.delete_rows(2, max_data_row-1)

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    fill1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Branco
    fill2 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Cinza claro

    # Escrever os dados das tarefas a partir da linha 2
    start_row = 2
    for idx, tarefa in enumerate(tarefas_processadas):
        row = start_row + idx
        ws.cell(row=row, column=1, value=tarefa['Código da Tarefa'])
        ws.cell(row=row, column=2, value=tarefa['Data'])
        ws.cell(row=row, column=3, value=tarefa['Cliente'])
        ws.cell(row=row, column=4, value=tarefa['Colaborador'])
        ws.cell(row=row, column=5, value=tarefa['Tipo de Tarefa'])
        ws.cell(row=row, column=6, value=', '.join(tarefa['Serviços']))
        ws.cell(row=row, column=7, value=tarefa['Faturamento'])
        ws.cell(row=row, column=8, value=tarefa['Lucro'])
        # Aplicar borda e cor alternada
        fill = fill1 if idx % 2 == 0 else fill2
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.fill = fill

    # Salvar em arquivo temporário
    temp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp.name)

    return temp.name
