from flask import Blueprint, request, jsonify, send_file
from app.controller.login.check_auth import check_auth
from app.models.user import User
from app.models.tarefas import Tarefas
from app.models.produtos import Produtos
from app.models.tipos_de_tarefas import Tipos_de_tarefas
import os
import tempfile
import openpyxl
from openpyxl.styles import Border, Side, PatternFill

planilha_produtos_bp = Blueprint('planilha_produtos', __name__, url_prefix='/gerar_planilha')

@planilha_produtos_bp.route('/produto', methods=['GET'])
def gerar_planilha_produto():
    api_key = request.args.get('api_key')
    auth = check_auth(api_key)

    if not auth[0]:
        return jsonify({"message": "user não autenticado"}), 401

    user = User.query.filter_by(api_key=api_key).first()
    user_id = user.id

    # Buscar dados das tarefas
    tarefas_obj = Tarefas.query.filter_by(user_id=user.id).first()
    tarefas = tarefas_obj.json_lista_tarefas if tarefas_obj and tarefas_obj.json_lista_tarefas else []

    # Filtrar apenas tarefas com faturamento de produtos diferente de zero
    tarefas_produtos = []
    for tarefa in tarefas:
        faturamento_produtos = tarefa.get('faturamento-produtos', 0)
        if faturamento_produtos and faturamento_produtos != 0:
            tarefas_produtos.append(tarefa)

    # Buscar produtos cadastrados
    produtos_obj = Produtos.query.filter_by(user_id=user.id).first()
    produtos_cadastrados = produtos_obj.json_lista_produtos if produtos_obj and produtos_obj.json_lista_produtos else []

    # Montar lista de tarefas com os campos desejados (apenas produtos)
    resultado = []
    for tarefa in tarefas_produtos:
        resultado.append({
            'Código da Tarefa': tarefa.get('id-da-tarefa'),
            'Data': tarefa.get('data-da-tarefa', '')[:10],
            'Cliente': tarefa.get('nome-do-cliente'),
            'Colaborador': tarefa.get('id-do-colaborador'),
            'Tipo de Tarefa': tarefa.get('tipo-da-tarefa'),
            'Produtos': tarefa.get('produtos', []),
            'Faturamento': tarefa.get('faturamento-produtos', 0),
            'Lucro': tarefa.get('lucro-produto', 0)
        })

    # Criar dicionário de id->nome para produtos cadastrados
    produtos_id_nome = {}
    for p in produtos_cadastrados:
        if isinstance(p, dict) and 'id-produto' in p and 'nome-do-produto' in p:
            produtos_id_nome[str(p['id-produto'])] = p['nome-do-produto']

    # Montar lista de tarefas com nomes de produtos
    tarefas_com_nomes = []
    for tarefa in resultado:
        nomes_produtos = [produtos_id_nome.get(str(pid), pid) for pid in tarefa['Produtos']]
        tarefa_nome = tarefa.copy()
        tarefa_nome['Produtos'] = nomes_produtos
        tarefas_com_nomes.append(tarefa_nome)

    # Buscar tipos de tarefa cadastrados
    tipos_obj = Tipos_de_tarefas.query.filter_by(user_id=user.id).first()
    tipos_cadastrados = tipos_obj.json_lista_tipos_de_tarefas if tipos_obj and tipos_obj.json_lista_tipos_de_tarefas else []

    # Criar dicionário de id->nome para tipos de tarefa cadastrados
    tipos_id_nome = {}
    for t in tipos_cadastrados:
        if isinstance(t, dict) and 'id-tipo-de-tarefa' in t and 'nome-do-tipo-de-tarefa' in t:
            tipos_id_nome[str(t['id-tipo-de-tarefa'])] = t['nome-do-tipo-de-tarefa']

    # Montar lista final de tarefas com nomes de produtos e tipos de tarefa
    tarefas_final = []
    for tarefa in tarefas_com_nomes:
        tipo_nome = tipos_id_nome.get(str(tarefa['Tipo de Tarefa']), tarefa['Tipo de Tarefa'])
        tarefa_nome = tarefa.copy()
        tarefa_nome['Tipo de Tarefa'] = tipo_nome
        tarefas_final.append(tarefa_nome)

    # Gerar planilha Excel usando o modelo de produtos
    modelo_path = os.path.join(os.path.dirname(__file__), 'modelos', 'Relatorio_de_Lucro_produto.xlsx')
    wb = openpyxl.load_workbook(modelo_path)
    ws = wb.active

    # Limpar linhas antigas de dados (exceto cabeçalho)
    max_data_row = ws.max_row
    if max_data_row > 2:
        ws.delete_rows(2, max_data_row-1)

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    fill1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Branco
    fill2 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Cinza claro

    # Escrever os dados das tarefas a partir da linha 2
    start_row = 2
    for idx, tarefa in enumerate(tarefas_final):
        row = start_row + idx
        ws.cell(row=row, column=1, value=tarefa['Código da Tarefa'])
        ws.cell(row=row, column=2, value=tarefa['Data'])
        ws.cell(row=row, column=3, value=tarefa['Cliente'])
        ws.cell(row=row, column=4, value=tarefa['Colaborador'])
        ws.cell(row=row, column=5, value=tarefa['Tipo de Tarefa'])
        ws.cell(row=row, column=6, value=', '.join(tarefa['Produtos']))
        ws.cell(row=row, column=7, value=tarefa['Faturamento'])
        ws.cell(row=row, column=8, value=tarefa['Lucro'])
        # Aplicar borda e cor alternada
        fill = fill1 if idx % 2 == 0 else fill2
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.fill = fill

    # Salvar em arquivo temporário
    temp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp.name)

    return send_file(temp.name, as_attachment=True, download_name='Relatorio_de_Lucro_Produto.xlsx'), 200
