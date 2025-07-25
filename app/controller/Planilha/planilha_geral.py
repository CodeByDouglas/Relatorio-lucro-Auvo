from flask import Blueprint, request, jsonify, send_file
from app.controller.login.check_auth import check_auth
from app.models.user import User
from app.models.tarefas import Tarefas
from app.models.produtos import Produtos
from app.models.servicos import Servicos
from app.models.tipos_de_tarefas import Tipos_de_tarefas
from app.models.dados_calculados import Faturamento_total, Lucro_total
import pandas as pd
import os
import tempfile
import openpyxl
from openpyxl.styles import Border, Side, PatternFill

planilha_bp = Blueprint('planilha', __name__, url_prefix='/gerar_planilha')

@planilha_bp.route('/geral', methods=['GET'])
def gerar_planilha_geral():
    api_key = request.args.get('api_key')
    auth = check_auth(api_key)

    if not auth[0]:
        return jsonify({"message": "user não autenticado"}), 401

    user = User.query.filter_by(api_key=api_key).first()
    user_id = user.id

    # Buscar dados das tarefas
    tarefas_obj = Tarefas.query.filter_by(user_id=user.id).first()
    tarefas = tarefas_obj.json_lista_tarefas if tarefas_obj and tarefas_obj.json_lista_tarefas else []

    # Buscar produtos cadastrados
    produtos_obj = Produtos.query.filter_by(user_id=user.id).first()
    produtos_cadastrados = produtos_obj.json_lista_produtos if produtos_obj and produtos_obj.json_lista_produtos else []

    # Montar lista de tarefas com os campos desejados
    resultado = []
    for tarefa in tarefas:
        resultado.append({
            'Código da Tarefa': tarefa.get('id-da-tarefa'),
            'Data': tarefa.get('data-da-tarefa', '')[:10],
            'Cliente': tarefa.get('nome-do-cliente'),
            'Colaborador': tarefa.get('id-do-colaborador'),
            'Tipo de Tarefa': tarefa.get('tipo-da-tarefa'),
            'Produtos': tarefa.get('produtos', []),
            'Serviços': tarefa.get('serviços', []),
            'Faturamento Produto': tarefa.get('faturamento-produtos', 0),
            'Faturamento Serviço': tarefa.get('faturamento-servicos', 0),
            'Lucro Produto': tarefa.get('lucro-produto', 0),
            'Lucro Serviço': tarefa.get('lucro-servicos', 0),
            'Faturamento': tarefa.get('faturamento-total', 0),
            'Lucro': tarefa.get('lucro-total', 0)
        })

    # Criar dicionário de id->nome para produtos cadastrados (ajustado para id-produto e nome-do-produto)
    produtos_id_nome = {}
    for p in produtos_cadastrados:
        if isinstance(p, dict) and 'id-produto' in p and 'nome-do-produto' in p:
            produtos_id_nome[str(p['id-produto'])] = p['nome-do-produto']

    # Montar lista de tarefas com nomes de produtos
    tarefas_com_nomes = []
    for tarefa in resultado:
        nomes_produtos = [produtos_id_nome.get(str(pid), pid) for pid in tarefa['Produtos']]
        tarefa_nome = tarefa.copy()
        tarefa_nome['Produtos'] = nomes_produtos
        tarefas_com_nomes.append(tarefa_nome)

    # Buscar tipos de tarefa cadastrados
    tipos_obj = Tipos_de_tarefas.query.filter_by(user_id=user.id).first()
    tipos_cadastrados = tipos_obj.json_lista_tipos_de_tarefas if tipos_obj and tipos_obj.json_lista_tipos_de_tarefas else []

    # Criar dicionário de id->nome para tipos de tarefa cadastrados
    tipos_id_nome = {}
    for t in tipos_cadastrados:
        if isinstance(t, dict) and 'id-tipo-de-tarefa' in t and 'nome-do-tipo-de-tarefa' in t:
            tipos_id_nome[str(t['id-tipo-de-tarefa'])] = t['nome-do-tipo-de-tarefa']

    # Montar lista de tarefas com nomes de produtos e tipos de tarefa
    tarefas_com_nomes_completos = []
    for tarefa in tarefas_com_nomes:
        tipo_nome = tipos_id_nome.get(str(tarefa['Tipo de Tarefa']), tarefa['Tipo de Tarefa'])
        tarefa_nome = tarefa.copy()
        tarefa_nome['Tipo de Tarefa'] = tipo_nome
        tarefas_com_nomes_completos.append(tarefa_nome)

    # Buscar serviços cadastrados
    servicos_obj = Servicos.query.filter_by(user_id=user.id).first()
    servicos_cadastrados = servicos_obj.json_lista_servicos if servicos_obj and servicos_obj.json_lista_servicos else []

    # Criar dicionário de id->nome para serviços cadastrados
    servicos_id_nome = {}
    for s in servicos_cadastrados:
        if isinstance(s, dict) and 'id-servico' in s and 'nome-do-servico' in s:
            servicos_id_nome[str(s['id-servico'])] = s['nome-do-servico']

    # Montar lista de tarefas com nomes de produtos, tipos de tarefa e serviços
    tarefas_com_nomes_tudo = []
    for tarefa in tarefas_com_nomes_completos:
        nomes_servicos = [servicos_id_nome.get(str(sid), sid) for sid in tarefa['Serviços']]
        tarefa_nome = tarefa.copy()
        tarefa_nome['Serviços'] = nomes_servicos
        tarefas_com_nomes_tudo.append(tarefa_nome)

    # Gerar planilha Excel usando o modelo e tarefas_com_nomes_tudo
    import openpyxl
    import tempfile
    import os
    modelo_path = os.path.join(os.path.dirname(__file__), 'modelos', 'Planilha_relatorio_de_lucro_geral.xlsx')
    wb = openpyxl.load_workbook(modelo_path)
    ws = wb.active

    # Limpar linhas antigas de dados (exceto cabeçalho)
    max_data_row = ws.max_row
    if max_data_row > 2:
        ws.delete_rows(2, max_data_row-1)

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    fill1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Branco
    fill2 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Cinza claro

    # Escrever os dados das tarefas a partir da linha 2
    start_row = 2
    for idx, tarefa in enumerate(tarefas_com_nomes_tudo):
        row = start_row + idx
        ws.cell(row=row, column=1, value=tarefa['Código da Tarefa'])
        ws.cell(row=row, column=2, value=tarefa['Data'])
        ws.cell(row=row, column=3, value=tarefa['Cliente'])
        ws.cell(row=row, column=4, value=tarefa['Colaborador'])
        ws.cell(row=row, column=5, value=tarefa['Tipo de Tarefa'])
        ws.cell(row=row, column=6, value=', '.join(tarefa['Produtos']))
        ws.cell(row=row, column=7, value=', '.join(tarefa['Serviços']))
        ws.cell(row=row, column=8, value=tarefa['Faturamento Produto'])
        ws.cell(row=row, column=9, value=tarefa['Faturamento Serviço'])
        ws.cell(row=row, column=10, value=tarefa['Lucro Produto'])
        ws.cell(row=row, column=11, value=tarefa['Lucro Serviço'])
        ws.cell(row=row, column=12, value=tarefa['Faturamento'])
        ws.cell(row=row, column=13, value=tarefa['Lucro'])
        # Aplicar borda e cor alternada
        fill = fill1 if idx % 2 == 0 else fill2
        for col in range(1, 14):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.fill = fill

    # Salvar em arquivo temporário
    temp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp.name)

    return send_file(temp.name, as_attachment=True, download_name='Relatorio_de_Lucro.xlsx'), 200

